# -*- coding: utf-8 -*-
#
import os
from collections import OrderedDict

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

from package.const import STATIC_PATH

TEMPLATE_PATH = os.path.join(STATIC_PATH, 'templates')


class PageFilter:
    def __init__(self):
        self.cur_page = 0

    def get_page(self, arg):
        self.cur_page += 1
        return self.cur_page


class ExcelPrinter:

    @staticmethod
    def __auto_adjust_column_width(ws):
        def calc_handler(x):
            return len(x.encode()) if x else 0

        pre_rows = ws.iter_cols(min_row=1, max_row=2, values_only=True)
        # 动态分配表头宽度
        column_width = [max(map(calc_handler, i)) for i in pre_rows]
        # 调整首行宽度
        for i, width in enumerate(column_width, 1):
            width = width if width < 100 else 100
            width = 10 if width < 10 else width
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = width

    @staticmethod
    def __fill_row_color(ws, row_numbers):
        fill = PatternFill(start_color='cccccc', end_color='cccccc', fill_type='solid')
        for n in row_numbers:
            for row in ws.iter_rows(min_row=n, max_row=n):
                for cell in row:
                    cell.fill = fill

    @staticmethod
    def _init_filter(env):
        env.filters['get_page'] = PageFilter().get_page

    @property
    def sheet1(self):
        """
        获取主机名、IP、系统用户名称、账号用户名、明文密码/密钥，该账号最近一次成功登录该资产的时间
        """
        return OrderedDict({
            'user_count': '活跃用户数量',
            'asset_count': '资产总数量',
            'sessions': '会话总数量',
            'organization_count': '组织数量',
            'linux_asset_count': 'Linux类型资产',
            'windows_asset_count': 'Windows类型资产',
            'database_asset_count': '数据库类型资产',
            'max_login_count': '最大单日登录次数',
            'max_connect_asset_count': '最大单日访问资产数',
            'last_1_month_login_count': '近一月登录用户数',
            'last_1_month_connect_asset_count': '近一月登录资产数',
            'last_1_month_upload_count': '近一月文件上传数',
            'last_3_month_connect_asset_count': '近三月登录资产数',
            'last_3_month_max_login_count': '近三月最大单日用户登录数',
            'last_3_month_max_connect_asset_count': '近三月最大单日资产登录数',
            'last_3_month_login_count': '近三月登录用户数',
            'last_3_month_upload_count': '近三月文件上传数',
            'last_3_month_command_count': '近三月命令记录数',
            'last_3_month_danger_command_count': '近三月高危命令记录数',
            'last_3_month_max_session_duration': '近三月最大会话时长',
            'last_3_month_avg_session_duration': '近三月平均会话时长(秒)',
            'last_3_month_ticket_count': '近三月工单申请数'
        })

    def save(self, filepath, context):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '巡检数据概览'
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['A1'] = 'JumpServer巡检数据报告'
        font_title = Font(size=14)
        ws['A1'].font = font_title
        for k, v in self.sheet1.items():
            ws.append([v, str(context['v'][k])])
        self.__auto_adjust_column_width(ws)

        ws = wb.create_sheet('资产类型分布前10')
        ws.append(['资产类型', '数量'])
        for item in context['v']['platform_chart']:
            ws.append([item['name'], str(item['value'])])

        ws = wb.create_sheet('会话调用组件类型数')
        ws.append(['组件', '调用组件数'])
        for item in context['v']['terminal_chart']:
            ws.append([item['name'], str(item['value'])])

        ws = wb.create_sheet('近3个月会话协议数')
        ws.append(['协议', '会话协议数'])
        for item in context['v']['protocol_chart']:
            ws.append([item['name'], str(item['value'])])

        ws = wb.create_sheet('近3个月用户登录数')
        ws.append(['日期', '用户登录数'])
        for index, item in enumerate(context['v']['user_login_chart']['x']):
            ws.append([item, str(context['v']['user_login_chart']['y'][index])])

        ws = wb.create_sheet('近3个月资产登录数')
        ws.append(['日期', '资产登录数'])
        for index, item in enumerate(context['v']['asset_connect_chart']['x']):
            ws.append([item, str(context['v']['asset_connect_chart']['y'][index])])

        ws = wb.create_sheet('近3个月活跃用户')
        ws.append(['用户', '登录数'])
        for index, item in enumerate(context['v']['active_user_chart']['x']):
            ws.append([item, str(context['v']['active_user_chart']['y'][index])])

        ws = wb.create_sheet('近3个月活跃资产')
        ws.append(['资产', '登录数'])
        for index, item in enumerate(context['v']['active_asset_chart']['x']):
            ws.append([item, str(context['v']['active_asset_chart']['y'][index])])

        ws = wb.create_sheet('功能使用情况')
        ws.append(['功能项', 'XPack', '是否启用'])
        for item in context['v']['settings_chart']:
            ws.append([item['name'], str(item['xpack']), str(item['value'])])
        self.__auto_adjust_column_width(ws)
        wb.save(filepath)
